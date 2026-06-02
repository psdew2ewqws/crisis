"""AEGIS Crisis — PROOF + Excel-report surface (proof.router APIRouter).

Two endpoints that COMPOSE the existing grounded engines (whys, rootcause,
cluster_link, the validate route, the forecast route) — they do NOT re-implement
any reasoning. Every number/label/quote is RETRIEVED from real voc360 rows.

  GET /api/proof?type=cluster|service|all&key=&depth=5
      → resolve the subject to a cluster, then assemble the full proof bundle
        (why-chain + root + narration + structured validation + representative
        evidence quotes + REAL the_data rows proving it + owning services +
        best-effort forecast) per the agreed API contract.

  GET /api/report/<cluster_id>.xlsx
      → stream a 4-sheet .xlsx (Summary, Why-Chain, Evidence, Related Cases),
        built in-memory with openpyxl. Arabic is preserved (xlsx is utf-8).

Every sub-call is wrapped in try/except so a slow/missing piece degrades to a
sensible default rather than a 500. ``main.py`` includes this with the same
``try/except include_router`` pattern it uses for ``main_v2`` / ``api_v3``.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

from . import db  # read-only voc360 access

# Real engines — import each defensively so one missing module never breaks import.
try:
    from . import whys
except Exception:  # pragma: no cover
    whys = None  # type: ignore

try:
    from . import rootcause
except Exception:  # pragma: no cover
    rootcause = None  # type: ignore

try:
    from . import cluster_link
except Exception:  # pragma: no cover
    cluster_link = None  # type: ignore

try:
    from . import api_v3  # validate_endpoint + forecast (existing routes/functions)
except Exception:  # pragma: no cover
    api_v3 = None  # type: ignore


router = APIRouter()

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _short_phrase(text: Any, words: int = 6) -> str:
    """Clip a raw cluster label (often a citizen sentence) to a short phrase."""
    t = str(text or "").strip()
    parts = t.split()
    return (" ".join(parts[:words]) + "…") if len(parts) > words else t


def _plain_summary(subject: Dict[str, Any], why: Dict[str, Any]) -> str:
    """A jargon-free, one-paragraph Arabic explanation for a NON-technical reader.

    No cluster ids, no confidence scores, no 'depth' — just: what citizens are
    complaining about, how big it is, which service, the deeper cause, and the
    suggested next step. This is the 'ببساطة' line shown at the top of the panel.
    """
    label = _short_phrase(subject.get("label_ar") or subject.get("label_en") or "هذه المشكلة")
    members = subject.get("members") or 0
    services = subject.get("services") or []
    svc = services[0][0] if services and isinstance(services[0], (list, tuple)) and services[0] else None
    # prefer the ARABIC root cause (last why-step's `because`); the top-level
    # `root` field is the English gloss, which we avoid in the plain-Arabic line.
    chain = (why or {}).get("why_chain") or []
    root_ar = ""
    if chain:
        last = chain[-1] or {}
        root_ar = str(last.get("because") or "").split("(")[0].strip()
    if not root_ar:
        root_ar = str((why or {}).get("root") or "").split("(")[0].strip()

    parts = [f"باختصار: يشتكي المواطنون بشأن «{label}»"]
    if members:
        parts.append(f"، وقد رصدنا {members} بلاغًا حول هذا الموضوع")
    if svc:
        parts.append(f"، أكثرها مرتبط بخدمة «{svc}»")
    parts.append(".")
    if root_ar and root_ar != label:
        parts.append(f" والسبب الأعمق وراء ذلك هو {root_ar}.")
    parts.append(" الخطوة المقترحة: إحالة الموضوع إلى الجهة المالكة لمعالجة جذره، ثم متابعة هل تنخفض الشكاوى بعد المعالجة.")
    return "".join(parts)


# =========================================================================== #
# Subject resolution + bundle assembly helpers (read-only, never raise).      #
# =========================================================================== #
def _resolve_cluster(stype: str, key: Optional[str]) -> Optional[str]:
    """Resolve the request subject to a cluster_id.

    - cluster: use ``key`` directly.
    - service: ``whys._dominant_cluster_for_service(key)``.
    - all: ``rootcause.rank_root_causes(1)[0]``.
    """
    stype = (stype or "all").lower()
    if stype == "cluster" and key:
        return key
    if stype == "service" and key and whys is not None:
        try:
            info = whys._dominant_cluster_for_service(key)
            if info and info.get("cluster_id"):
                return info["cluster_id"]
        except Exception:
            pass
    if rootcause is not None:
        try:
            ranked = rootcause.rank_root_causes(1)
            if ranked:
                return ranked[0]["cluster_id"]
        except Exception:
            pass
    return None


def _cluster_row(cluster_id: str) -> Optional[Dict[str, Any]]:
    """Real ril_problem_clusters row (label/members/severity/first&last seen)."""
    if not cluster_id:
        return None
    try:
        return db.fetchone(
            """
            select cluster_id, canonical_label_ar, canonical_label_en,
                   coalesce(member_count, 0) as member_count,
                   coalesce(severity_avg, 0) as severity_avg,
                   first_seen, last_seen
            from ril_problem_clusters
            where cluster_id = %(cid)s
            """,
            {"cid": cluster_id},
        )
    except Exception:
        return None


def _build_subject(stype: str, key: Optional[str], cluster_id: str) -> Dict[str, Any]:
    crow = _cluster_row(cluster_id) or {}
    services: List[Any] = []
    signals = 0
    if cluster_link is not None:
        try:
            services = [list(x) for x in cluster_link.cluster_services(cluster_id)]
        except Exception:
            services = []
        try:
            signals = int(cluster_link.cluster_signals(cluster_id))
        except Exception:
            signals = 0
    return {
        "type": stype,
        "key": key,
        "cluster_id": cluster_id,
        "label_ar": crow.get("canonical_label_ar"),
        "label_en": crow.get("canonical_label_en"),
        "members": int(crow.get("member_count") or 0),
        "severity_avg": round(float(crow.get("severity_avg") or 0.0), 2),
        "signals": signals,
        "services": services,
        "first_seen": crow.get("first_seen"),
        "last_seen": crow.get("last_seen"),
    }


def _why_bundle(stype: str, key: Optional[str], cluster_id: str, depth: int) -> Dict[str, Any]:
    """why_chain + root + narration from whys.ask_whys (degrade to empties)."""
    if whys is None:
        return {"why_chain": [], "root": "", "narration": ""}
    # For a service start, ask from the service so depth-1 reads symptom→cluster;
    # otherwise anchor on the resolved cluster.
    if stype == "service" and key:
        start = {"type": "service", "key": key}
    else:
        start = {"type": "cluster", "key": cluster_id}
    try:
        out = whys.ask_whys(start, max_depth=depth)
    except Exception:
        return {"why_chain": [], "root": "", "narration": ""}
    chain = []
    for st in (out.get("chain") or []):
        chain.append({
            "depth": st.get("depth"),
            "question": st.get("question"),
            "because": st.get("because"),
            "because_en": st.get("because_en"),
            "evidence": [e for e in (st.get("evidence") or [])][:3],
            "signals": st.get("signals"),
        })
    root = out.get("root")
    root_str = ""
    if isinstance(root, dict):
        root_str = root.get("because_en") or root.get("because") or ""
    elif root:
        root_str = str(root)
    return {"why_chain": chain, "root": root_str, "narration": out.get("narration") or ""}


def _validation(cluster_id: str, service: Optional[str]) -> Dict[str, Any]:
    """Structured proof via the existing validate function."""
    if api_v3 is None:
        return {"verdict": "insufficient", "confidence": 0.0, "score": 0,
                "summary": "validate engine unavailable", "checks": []}
    try:
        return api_v3._run_validate(cluster_id, service)
    except Exception:
        return {"verdict": "insufficient", "confidence": 0.0, "score": 0,
                "summary": "validation failed", "checks": []}


def _evidence_segments(cluster_id: str, limit: int = 10) -> List[Dict[str, Any]]:
    """Top representative member quotes via whys._fetch_member_segments."""
    if whys is None:
        return []
    try:
        rows = whys._fetch_member_segments(cluster_id, limit=limit)
    except Exception:
        return []
    out = []
    for r in rows[:limit]:
        out.append({
            "segment_text": r.get("segment_text"),
            "confidence": r.get("confidence"),
        })
    return out


def _related_cases(cluster_id: str, limit: int = 50) -> List[Dict[str, Any]]:
    """REAL the_data rows proving the cluster, via cluster_link.cluster_signal_rows."""
    if cluster_link is None:
        return []
    try:
        rows = cluster_link.cluster_signal_rows(cluster_id, limit=limit)
    except Exception:
        return []
    out = []
    for r in rows:
        out.append({
            "record_id": r.get("record_id"),
            "service_id": r.get("service_id"),
            "text": r.get("text"),
            "sentiment_label": r.get("sentiment_label"),
            "severity": r.get("severity"),
            "observed_at": r.get("observed_at"),
            "source_type": r.get("source_type"),
        })
    return out


def _forecast(stype: str, key: Optional[str], cluster_id: str) -> Optional[Dict[str, Any]]:
    """Best-effort forecast via the existing forecast function; null on any error."""
    if api_v3 is None:
        return None
    entity = stype if stype in ("service", "cluster") else "cluster"
    fkey = key if (stype == "service" and key) else cluster_id
    try:
        fc = api_v3.forecast(entity=entity, key=fkey, metric="volume", horizon=30)
    except Exception:
        return None
    if not isinstance(fc, dict) or not fc.get("ok"):
        return None
    return {
        "history": fc.get("history", []),
        "forecast": fc.get("forecast", []),
        "escalation": fc.get("escalation"),
        "source": fc.get("source"),
    }


# =========================================================================== #
# 1) PROOF                                                                     #
# =========================================================================== #
@router.get("/api/proof")
def proof(
    type: str = Query("all", pattern="^(service|cluster|all)$"),
    key: Optional[str] = Query(default=None),
    depth: int = Query(5, ge=1, le=5),
) -> Dict[str, Any]:
    """Assemble the full grounded proof bundle for a service / cluster / national.

    Composes whys + validate + evidence + REAL the_data rows + services +
    best-effort forecast. Each sub-call degrades to a default so a slow/missing
    piece can't 500 the response.
    """
    cluster_id = _resolve_cluster(type, key)
    if not cluster_id:
        return {
            "ok": False,
            "error": "no cluster could be resolved for this subject",
            "subject": {"type": type, "key": key, "cluster_id": None},
            "why_chain": [], "root": "", "narration": "",
            "validation": {"verdict": "insufficient", "confidence": 0.0, "score": 0,
                           "summary": "no cluster resolved", "checks": []},
            "evidence_segments": [], "related_cases": [], "forecast": None,
            "report_url": None,
        }

    service = key if type == "service" else None
    subject = _build_subject(type, key, cluster_id)
    why = _why_bundle(type, key, cluster_id, depth)
    validation = _validation(cluster_id, service)
    evidence_segments = _evidence_segments(cluster_id, limit=10)
    related_cases = _related_cases(cluster_id, limit=50)
    services = []
    if cluster_link is not None:
        try:
            services = [list(x) for x in cluster_link.cluster_services(cluster_id)]
        except Exception:
            services = []
    forecast = _forecast(type, key, cluster_id)

    return {
        "ok": True,
        "subject": subject,
        "plain": _plain_summary(subject, why),
        "why_chain": why["why_chain"],
        "root": why["root"],
        "narration": why["narration"],
        "validation": validation,
        "evidence_segments": evidence_segments,
        "related_cases": related_cases,
        "services": services,
        "forecast": forecast,
        "report_url": f"/api/report/{cluster_id}.xlsx",
    }


# =========================================================================== #
# 2) REPORT — streamed .xlsx                                                   #
# =========================================================================== #
def _cell(v: Any) -> Any:
    """Coerce a DB value to an xlsx-safe cell (datetimes/None → str)."""
    if v is None:
        return ""
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _build_workbook(cluster_id: str) -> "io.BytesIO":
    from openpyxl import Workbook

    wb = Workbook()
    crow = _cluster_row(cluster_id)

    ws = wb.active
    ws.title = "Summary"

    if not crow:
        ws.append(["cluster_id", cluster_id])
        ws.append(["status", "cluster not found"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # --- Summary -----------------------------------------------------------
    services = []
    signals = 0
    if cluster_link is not None:
        try:
            services = [s for s, _ in cluster_link.cluster_services(cluster_id)]
        except Exception:
            services = []
        try:
            signals = int(cluster_link.cluster_signals(cluster_id))
        except Exception:
            signals = 0
    score = 0
    try:
        if rootcause is not None:
            for rc in rootcause.rank_root_causes(50):
                if rc["cluster_id"] == cluster_id:
                    score = rc.get("score") or 0
                    break
    except Exception:
        score = 0

    ws.append(["field", "value"])
    ws.append(["cluster_id", _cell(cluster_id)])
    ws.append(["label_ar", _cell(crow.get("canonical_label_ar"))])
    ws.append(["label_en", _cell(crow.get("canonical_label_en"))])
    ws.append(["severity_avg", round(float(crow.get("severity_avg") or 0.0), 2)])
    ws.append(["members", int(crow.get("member_count") or 0)])
    ws.append(["score", _cell(score)])
    ws.append(["signals", signals])
    ws.append(["services", ", ".join(_cell(s) for s in services)])
    ws.append(["first_seen", _cell(crow.get("first_seen"))])
    ws.append(["last_seen", _cell(crow.get("last_seen"))])

    # --- Why-Chain ---------------------------------------------------------
    ws_w = wb.create_sheet("Why-Chain")
    ws_w.append(["depth", "question", "because_en", "because", "evidence"])
    why = _why_bundle("cluster", None, cluster_id, 5)
    for st in why["why_chain"]:
        ev = " | ".join(_cell(e) for e in (st.get("evidence") or [])[:3])
        ws_w.append([
            _cell(st.get("depth")),
            _cell(st.get("question")),
            _cell(st.get("because_en")),
            _cell(st.get("because")),
            ev,
        ])

    # --- Evidence ----------------------------------------------------------
    ws_e = wb.create_sheet("Evidence")
    ws_e.append(["segment_text", "confidence"])
    for seg in _evidence_segments(cluster_id, limit=20):
        ws_e.append([_cell(seg.get("segment_text")), _cell(seg.get("confidence"))])

    # --- Related Cases -----------------------------------------------------
    ws_r = wb.create_sheet("Related Cases")
    ws_r.append(["record_id", "service_id", "text", "sentiment_label", "severity", "observed_at"])
    for rc in _related_cases(cluster_id, limit=60):
        ws_r.append([
            _cell(rc.get("record_id")),
            _cell(rc.get("service_id")),
            _cell(rc.get("text")),
            _cell(rc.get("sentiment_label")),
            _cell(rc.get("severity")),
            _cell(rc.get("observed_at")),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/api/report/{cluster_id}.xlsx")
def report(cluster_id: str) -> StreamingResponse:
    """Stream a 4-sheet .xlsx evidence report for a cluster (in-memory)."""
    buf = _build_workbook(cluster_id)
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={
            "Content-Disposition": f"attachment; filename=crisis-proof-{cluster_id[:8]}.xlsx"
        },
    )


__all__ = ["router"]
